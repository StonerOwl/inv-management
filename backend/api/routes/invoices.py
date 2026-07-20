"""
Invoice CRUD routes + search + CSV/Excel export.
"""
import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_

from db.database import get_db
from db.repository import InvoiceRepository
from api.dependencies import get_current_admin
from db.models import InvoiceProjectAssignment, PWSItem
from pydantic import BaseModel
import uuid

router = APIRouter(prefix="/api/invoices", tags=["invoices"])

class ManualInvoiceCreate(BaseModel):
    invoice_number: str
    invoice_date: str
    order_id: Optional[str] = None
    seller_gstin: Optional[str] = None
    product_description: str
    quantity: float
    grand_total: float
    category: Optional[str] = "Category 1"


@router.get("")
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    repo = InvoiceRepository(db)
    invoices, total = repo.list_invoices(
        skip=skip, limit=limit, status=status, search=search
    )

    # Batch-fetch all project assignments for this page of invoices (avoids N+1).
    invoice_ids = [inv.id for inv in invoices]
    assignments = (
        db.query(InvoiceProjectAssignment)
        .options(joinedload(InvoiceProjectAssignment.group))
        .filter(InvoiceProjectAssignment.invoice_id.in_(invoice_ids))
        .all()
    ) if invoice_ids else []

    # Fetch the PWSItem rows needed for project_code + name
    project_ids_needed = list({a.project_id for a in assignments})
    pws_items = (
        db.query(PWSItem)
        .filter(PWSItem.id.in_(project_ids_needed))
        .all()
    ) if project_ids_needed else []
    pws_map = {p.id: p for p in pws_items}

    # Group assignments by invoice_id
    from collections import defaultdict
    assign_by_invoice = defaultdict(list)
    for a in assignments:
        pws = pws_map.get(a.project_id)
        assign_by_invoice[a.invoice_id].append({
            "project_id": a.project_id,
            "project_code": pws.project_code if pws else None,
            "project_name": pws.name if pws else None,
            "group_id": a.group_id,
            "group_name": a.group.name if a.group else None,
            "group_color": a.group.color if a.group else None,
        })

    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [
            {
                **inv.to_dict(),
                "line_items": [li.to_dict() for li in inv.line_items],
                "taxes": [t.to_dict() for t in inv.taxes],
                "project_assignments": assign_by_invoice[inv.id],
            }
            for inv in invoices
        ],
    }

@router.get("/advanced-search")
def advanced_search_invoices(
    query: Optional[str] = None,
    invoice_category: Optional[str] = None,
    po_category: Optional[str] = None,
    item_code: Optional[str] = None,
    process_name: Optional[str] = None,
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db)
):
    from db.models import Invoice, LineItem, InvoiceProcessTracking, WorkflowProcess
    
    q = db.query(Invoice)\
        .outerjoin(LineItem, Invoice.id == LineItem.invoice_id)\
        .outerjoin(InvoiceProcessTracking, Invoice.id == InvoiceProcessTracking.invoice_id)\
        .outerjoin(WorkflowProcess, InvoiceProcessTracking.process_id == WorkflowProcess.id)
    
    if query:
        pattern = f"%{query}%"
        q = q.filter(
            or_(
                Invoice.invoice_number.ilike(pattern),
                Invoice.order_id.ilike(pattern),
                Invoice.seller_gstin.ilike(pattern),
                LineItem.name.ilike(pattern)
            )
        )
        
    if invoice_category:
        q = q.filter(Invoice.category.ilike(f"%{invoice_category}%"))
    if process_name:
        q = q.filter(WorkflowProcess.name.ilike(f"%{process_name}%"))
    if status:
        q = q.filter(Invoice.status.ilike(f"%{status}%"))
        
    q = q.distinct()
    
    total = q.count()
    invoices = q.order_by(Invoice.created_at.desc()).offset(skip).limit(limit).all()
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [
            {
                **inv.to_dict(),
                "line_items": [li.to_dict() for li in inv.line_items],
                "taxes": [t.to_dict() for t in inv.taxes],
                "tracking": [trk.to_dict() for trk in db.query(InvoiceProcessTracking).filter(InvoiceProcessTracking.invoice_id == inv.id).all()]
            }
            for inv in invoices
        ]
    }


@router.post("/manual")
def create_manual_invoice(
    data: ManualInvoiceCreate,
    db: Session = Depends(get_db)
):
    from db.models import Invoice, LineItem
    # Create the manual invoice
    new_invoice = Invoice(
        file_name="Manual Entry",
        file_path="manual",
        file_hash=uuid.uuid4().hex,
        source_type="manual",
        invoice_number=data.invoice_number,
        invoice_date=data.invoice_date,
        order_id=data.order_id,
        seller_gstin=data.seller_gstin,
        grand_total=data.grand_total,
        category=data.category,
        confidence_score=100.0,
        status="processed"
    )
    db.add(new_invoice)
    db.commit()
    db.refresh(new_invoice)
    
    # Add a single line item for the product description and quantity
    line_item = LineItem(
        invoice_id=new_invoice.id,
        name=data.product_description,
        quantity=data.quantity,
        total_price=data.grand_total
    )
    db.add(line_item)
    db.commit()
    db.refresh(new_invoice)
    
    return new_invoice.to_dict()


# ── Export routes MUST be declared before /{invoice_id} to avoid
#    FastAPI matching "export" as an invoice_id path parameter. ────────────────

@router.get("/export/csv")
def export_csv(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    repo = InvoiceRepository(db)
    invoices = repo.get_all_for_export(status=status)

    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "ID", "Invoice Number", "Invoice Date",
        "Order ID", "Seller GSTIN",
        "Product Description", "HSN Code",
        "CGST Rate", "CGST Amount",
        "SGST Rate", "SGST Amount",
        "IGST Rate", "IGST Amount",
        "Total Tax", "Grand Total",
        "Confidence Score", "Status",
    ]
    writer.writerow(headers)

    for inv in invoices:
        # Flatten line items into a single string
        product_desc = "; ".join(
            li.name for li in inv.line_items if li.name
        ) or ""
        hsn_codes = "; ".join(
            li.hsn_code for li in inv.line_items if li.hsn_code
        ) or ""

        # Build tax lookup
        tax_map = {t.tax_type: t for t in inv.taxes}
        cgst = tax_map.get("CGST")
        sgst = tax_map.get("SGST")
        igst = tax_map.get("IGST")
        total_tax = sum(t.amount or 0 for t in inv.taxes)

        writer.writerow([
            inv.id, inv.invoice_number,
            inv.invoice_date, inv.order_id, inv.seller_gstin,
            product_desc, hsn_codes,
            cgst.rate if cgst else "", cgst.amount if cgst else "",
            sgst.rate if sgst else "", sgst.amount if sgst else "",
            igst.rate if igst else "", igst.amount if igst else "",
            total_tax if total_tax else "",
            inv.grand_total, inv.confidence_score, inv.status,
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=invoices.csv"},
    )


@router.get("/export/excel")
def export_excel(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    repo = InvoiceRepository(db)
    invoices = repo.get_all_for_export(status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "ID", "Invoice Number", "Invoice Date",
        "Order ID", "Seller GSTIN",
        "Product Description", "HSN Code",
        "CGST Rate", "CGST Amount",
        "SGST Rate", "SGST Amount",
        "IGST Rate", "IGST Amount",
        "Total Tax", "Grand Total",
        "Confidence Score", "Status",
    ]

    header_fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for inv in invoices:
        # Flatten line items
        product_desc = "; ".join(
            li.name for li in inv.line_items if li.name
        ) or ""
        hsn_codes = "; ".join(
            li.hsn_code for li in inv.line_items if li.hsn_code
        ) or ""

        # Build tax lookup
        tax_map = {t.tax_type: t for t in inv.taxes}
        cgst = tax_map.get("CGST")
        sgst = tax_map.get("SGST")
        igst = tax_map.get("IGST")
        total_tax = sum(t.amount or 0 for t in inv.taxes)

        ws.append([
            inv.id, inv.invoice_number,
            inv.invoice_date, inv.order_id, inv.seller_gstin,
            product_desc, hsn_codes,
            cgst.rate if cgst else None, cgst.amount if cgst else None,
            sgst.rate if sgst else None, sgst.amount if sgst else None,
            igst.rate if igst else None, igst.amount if igst else None,
            total_tax if total_tax else None,
            inv.grand_total, inv.confidence_score, inv.status,
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices.xlsx"},
    )


# ── Parameterized routes (must come after literal paths like /export/*) ──────

@router.get("/{invoice_id}")
def get_invoice(invoice_id: int, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    inv = repo.get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {
        **inv.to_dict(),
        "raw_text": inv.raw_text,
        "line_items": [li.to_dict() for li in inv.line_items],
        "taxes": [t.to_dict() for t in inv.taxes],
    }


@router.put("/{invoice_id}")
def update_invoice(invoice_id: int, updates: dict, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    inv = repo.update_invoice(invoice_id, updates)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv.to_dict()


@router.delete("/{invoice_id}", dependencies=[Depends(get_current_admin)])
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    ok = repo.delete_invoice(invoice_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Deleted"}