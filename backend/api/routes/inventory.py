"""
Inventory items CRUD routes.
Items are created automatically when invoices are registered to projects,
and can be viewed, edited, and deleted independently.
"""
from typing import Optional, Dict, Any, List

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, or_
from pydantic import BaseModel

from db.database import get_db
from db.models import InventoryItem, InvoiceProjectAssignment, PWSItem

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


class InventoryItemUpdate(BaseModel):
    item_name: Optional[str] = None
    notes: Optional[str] = None
    quantity: Optional[float] = None
    unit_price: Optional[float] = None
    total_amount: Optional[float] = None


@router.get("/items", response_model=Dict[str, Any])
def list_inventory_items(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """List all inventory items with optional search and pagination."""
    q = db.query(InventoryItem)

    if search:
        pattern = f"%{search}%"
        q = q.filter(
            or_(
                InventoryItem.item_name.ilike(pattern),
                InventoryItem.invoice_number.ilike(pattern),
                InventoryItem.source_file_name.ilike(pattern),
                InventoryItem.notes.ilike(pattern),
            )
        )
        
    if group_id:
        q = q.join(InvoiceProjectAssignment, InventoryItem.invoice_id == InvoiceProjectAssignment.invoice_id)\
             .filter(InvoiceProjectAssignment.group_id == group_id)

    total = q.count()
    items = q.order_by(desc(InventoryItem.created_at)).offset(skip).limit(limit).all()

    # Batch-fetch project assignments for these invoice IDs (2 queries, no N+1)
    invoice_ids = list({item.invoice_id for item in items})
    assignments = (
        db.query(InvoiceProjectAssignment)
        .options(joinedload(InvoiceProjectAssignment.group))
        .filter(InvoiceProjectAssignment.invoice_id.in_(invoice_ids))
        .all()
    ) if invoice_ids else []

    project_ids_needed = list({a.project_id for a in assignments})
    pws_items = (
        db.query(PWSItem)
        .filter(PWSItem.id.in_(project_ids_needed))
        .all()
    ) if project_ids_needed else []
    pws_map = {p.id: p for p in pws_items}

    from collections import defaultdict
    projects_by_invoice = defaultdict(list)
    for a in assignments:
        pws = pws_map.get(a.project_id)
        projects_by_invoice[a.invoice_id].append({
            "project_id": a.project_id,
            "project_code": pws.project_code if pws else None,
            "project_name": pws.name if pws else None,
            "group_id": a.group_id,
            "group_name": a.group.name if a.group else None,
            "group_color": a.group.color if a.group else None,
        })

    def item_with_project(item):
        d = item.to_dict()
        d["project_assignments"] = projects_by_invoice[item.invoice_id]
        return d

    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [item_with_project(item) for item in items],
    }


@router.get("/items/{item_id}", response_model=Dict[str, Any])
@router.get("/items/{item_id}", response_model=Dict[str, Any])
def get_inventory_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    assignments = db.query(InvoiceProjectAssignment)\
        .options(joinedload(InvoiceProjectAssignment.group))\
        .filter(InvoiceProjectAssignment.invoice_id == item.invoice_id).all()
        
    pws_ids = list({a.project_id for a in assignments})
    pws_items = db.query(PWSItem).filter(PWSItem.id.in_(pws_ids)).all() if pws_ids else []
    pws_map = {p.id: p for p in pws_items}

    proj_list = []
    for a in assignments:
        pws = pws_map.get(a.project_id)
        proj_list.append({
            "project_id": a.project_id,
            "project_code": pws.project_code if pws else None,
            "project_name": pws.name if pws else None,
            "group_id": a.group_id,
            "group_name": a.group.name if a.group else None,
            "group_color": a.group.color if a.group else None,
        })
        
    d = item.to_dict()
    d["project_assignments"] = proj_list
    return d


@router.get("/export/excel")
def export_inventory_excel(
    search: Optional[str] = None,
    group_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse

    q = db.query(InventoryItem)

    if search:
        pattern = f"%{search}%"
        q = q.filter(
            or_(
                InventoryItem.item_name.ilike(pattern),
                InventoryItem.invoice_number.ilike(pattern),
                InventoryItem.source_file_name.ilike(pattern),
                InventoryItem.notes.ilike(pattern),
            )
        )
        
    if group_id:
        q = q.join(InvoiceProjectAssignment, InventoryItem.invoice_id == InvoiceProjectAssignment.invoice_id)\
             .filter(InvoiceProjectAssignment.group_id == group_id)

    items = q.order_by(desc(InventoryItem.created_at)).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Items"

    headers = [
        "ID", "Item Name", "Invoice Number", "Quantity", "Unit Price", "Total Amount", 
        "HSN Code", "Tax Type", "Tax Amount", "Notes", "Source File"
    ]
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=item.item_name)
        ws.cell(row=row_num, column=3, value=item.invoice_number)
        ws.cell(row=row_num, column=4, value=item.quantity)
        ws.cell(row=row_num, column=5, value=item.unit_price)
        ws.cell(row=row_num, column=6, value=item.total_amount)
        ws.cell(row=row_num, column=7, value=item.hsn_code)
        ws.cell(row=row_num, column=8, value=item.tax_type)
        ws.cell(row=row_num, column=9, value=item.tax_amount)
        ws.cell(row=row_num, column=10, value=item.notes)
        ws.cell(row=row_num, column=11, value=item.source_file_name)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_items.xlsx"}
    )
    """Get a single inventory item by ID."""
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    return item.to_dict()


@router.put("/items/{item_id}", response_model=Dict[str, Any])
def update_inventory_item(
    item_id: int,
    update: InventoryItemUpdate,
    db: Session = Depends(get_db),
):
    """Update an inventory item's name, notes, or quantities."""
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")

    if update.item_name is not None:
        item.item_name = update.item_name
    if update.notes is not None:
        item.notes = update.notes
    if update.quantity is not None:
        item.quantity = update.quantity
    if update.unit_price is not None:
        item.unit_price = update.unit_price
    if update.total_amount is not None:
        item.total_amount = update.total_amount

    db.commit()
    db.refresh(item)
    return item.to_dict()


@router.delete("/items/{item_id}")
def delete_inventory_item(item_id: int, db: Session = Depends(get_db)):
    """Delete an inventory item."""
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")

    db.delete(item)
    db.commit()
    return {"message": "Inventory item deleted"}